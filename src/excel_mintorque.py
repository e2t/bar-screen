import os
from datetime import datetime

from dry.measurements import to_mm
from openpyxl import Workbook
from openpyxl.cell import Cell
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from calc import (DEFAULT_TILT_ANGLE, DRIVE_UNITS_BIG, DRIVE_UNITS_SMALL,
                  FILTERS, HSDATA, NOMINAL_GAPS, WSDATA, BarScreen, InputData)

THIN = Side(border_style='thin', color='000000')

DRIVES = list(DRIVE_UNITS_SMALL.keys()) + list(DRIVE_UNITS_BIG.keys())
COLORS = {
    DRIVES[0]: 'ffc7ce',
    DRIVES[1]: 'ffeb9c',
    DRIVES[2]: 'c6efce',
    DRIVES[3]: '00b0f0',
    DRIVES[4]: 'ffff00',
}
COLUMNS = [chr(i) for i in range(ord('A'), ord('Z') + 1)]
COLUMNS += ['A' + chr(i) for i in range(ord('A'), ord('Z') + 1)]


def addborder(cell: Cell,
              top: Side | None = None,
              bottom: Side | None = None,
              left: Side | None = None,
              right: Side | None = None) -> None:
    old = cell.border
    cell.border = Border(top=top or old.top,
                         bottom=bottom or old.bottom,
                         left=left or old.left,
                         right=right or old.right)


def rectangle(cell_range: tuple[tuple[Cell]] | Cell) -> None:
    if isinstance(cell_range, Cell):
        addborder(cell_range, top=THIN, bottom=THIN, left=THIN, right=THIN)
        return
    for col, i in enumerate(cell_range[0]):
        addborder(i, top=THIN)
        addborder(cell_range[-1][col], bottom=THIN)
    for row in cell_range:
        addborder(row[0], left=THIN)
        addborder(row[-1], right=THIN)


def fillcolor(cell_range: tuple[tuple[Cell]] | Cell, color: str) -> None:
    if isinstance(cell_range, Cell):
        cell_range.fill = PatternFill('solid', fgColor=color)
        return
    for row in cell_range:
        for i in row:
            i.fill = PatternFill('solid', fgColor=color)


def main() -> None:
    book = Workbook()
    sheet = book.active

    row: int
    for hi, hs in enumerate(reversed(HSDATA)):
        row = 2 + hi
        sheet.cell(row=row, column=1, value=f'xx{hs:02d}')
        for wi, ws in enumerate(WSDATA):
            inp = InputData(width=(ws + 1) * 0.1,
                            depth=hs * 0.1,
                            mindrop=0.1,
                            gap=NOMINAL_GAPS[0],
                            ws=ws,
                            hs=hs,
                            gs=hs,
                            fp=FILTERS[0],
                            flow=None,
                            level=None,
                            angle=DEFAULT_TILT_ANGLE)
            scr = BarScreen(inp, None)
            cell = sheet.cell(row=row, column=2 + wi)
            cell.alignment = Alignment(horizontal='right')
            cell.value = f'{round(scr.mintorque)}, '\
                         f'{to_mm(scr.spring_preload):n}'
            if scr.drive:
                fillcolor(cell, COLORS[scr.drive])
                if not scr.is_drive_passed:
                    cell.font = Font(underline='single')

    row += 1
    for wi, ws in enumerate(WSDATA):
        cell = sheet.cell(row=row, column=2 + wi, value=f'{ws:02d}xx')
        cell.alignment = Alignment(horizontal='right')
        letter = COLUMNS[1 + wi]
        rectangle(sheet[f'{letter}2:{letter}59'])

    title = sheet.cell(row=1, column=1,
                       value='Необхідний обертовий момент РКЕ (Нм) та '
                             'попередній стиск пружини (мм)')
    title.font = Font(b=True)
    title.alignment = Alignment(horizontal='center')
    sheet.merge_cells(start_row=1, start_column=1,
                      end_row=1, end_column=len(WSDATA) + 1)

    rectangle(sheet['A2:AA59'])
    rectangle(sheet['A2:A59'])
    rectangle(sheet['A59:AA59'])

    row += 2
    for i, (drive, color) in enumerate(COLORS.items()):
        cell = sheet.cell(row=row + i, column=2,
                          value=f'{drive.designation}; {drive.torque:n} Нм')
        fillcolor(sheet.cell(row=row + i, column=1), color)

    for i in range(1, len(WSDATA) + 2):
        sheet.column_dimensions[get_column_letter(i)].width = 10

    today = datetime.now().strftime('%Y-%m-%d')
    filename = f'RKE gearmotors ({today}).xlsx'
    book.save(filename)
    os.system(f'"{filename}"')


if __name__ == '__main__':
    main()
